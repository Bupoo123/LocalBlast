#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
更新参比序列数据库
从Excel文件读取数据并更新species_db.json
"""

import json
import re
import sys
import os

try:
    import openpyxl
except ImportError:
    print("错误: 需要安装openpyxl库")
    print("请运行: pip3 install openpyxl")
    sys.exit(1)

def clean_sequence(seq):
    """清理序列，只保留ATCG字符"""
    if not seq:
        return ""
    # 转换为大写，移除空白字符
    seq = str(seq).upper().replace(" ", "").replace("\n", "").replace("\t", "")
    # 只保留ATCG和可能的模糊字符，最后只保留ATCG
    seq = re.sub(r'[^ATCGNMRWSYKVHDBNX-]', '', seq)
    # 移除所有非标准字符，只保留ATCG
    seq = re.sub(r'[^ATCG]', '', seq)
    return seq

def read_excel_to_species_db(excel_file):
    """从Excel文件读取数据并转换为species_db格式"""
    print(f"正在读取Excel文件: {excel_file}")
    
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active
        
        print(f"工作表名称: {ws.title}")
        print(f"总行数: {ws.max_row}")
        print(f"总列数: {ws.max_column}")
        
        # 读取表头
        headers = []
        header_row = 1
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col).value
            if cell_value:
                headers.append(str(cell_value).strip())
            else:
                headers.append(f"列{col}")
        
        print(f"表头: {headers}")
        
        # 尝试识别列名（支持多种可能的列名）
        name_col = None
        code_col = None
        sequence_col = None
        id_col = None
        
        for i, header in enumerate(headers):
            header_lower = header.lower()
            if '名称' in header or 'name' in header_lower or '物种' in header or '靶点' in header:
                name_col = i + 1
            elif '编号' in header or 'code' in header_lower or '代码' in header:
                code_col = i + 1
            elif ('参比序列' in header or 'sequence' in header_lower or ('seq' in header_lower and '大小' not in header)):
                sequence_col = i + 1
            elif 'id' in header_lower or '序号' in header or '流水号' in header:
                id_col = i + 1
        
        print(f"识别到的列位置:")
        print(f"  ID列: {id_col} ({headers[id_col-1] if id_col else '未找到'})")
        print(f"  名称列: {name_col} ({headers[name_col-1] if name_col else '未找到'})")
        print(f"  编号列: {code_col} ({headers[code_col-1] if code_col else '未找到'})")
        print(f"  序列列: {sequence_col} ({headers[sequence_col-1] if sequence_col else '未找到'})")
        
        if not name_col or not sequence_col:
            print("错误: 无法识别必要的列（名称和序列）")
            print("请确保Excel文件包含以下列之一:")
            print("  - 名称/Name/物种")
            print("  - 序列/Sequence/Seq")
            return None
        
        # 读取数据
        species_list = []
        max_id = 0
        
        for row in range(2, ws.max_row + 1):  # 从第2行开始（跳过表头）
            name = ws.cell(row=row, column=name_col).value
            sequence = ws.cell(row=row, column=sequence_col).value if sequence_col else None
            code = ws.cell(row=row, column=code_col).value if code_col else ""
            id_value = ws.cell(row=row, column=id_col).value if id_col else None
            
            # 跳过空行
            if not name or not sequence:
                continue
            
            name = str(name).strip()
            sequence = clean_sequence(sequence)
            code = str(code).strip() if code else ""
            
            if not sequence or len(sequence) < 10:
                print(f"警告: 第{row}行的序列太短或无效，已跳过")
                continue
            
            # 处理ID
            if id_value:
                try:
                    species_id = int(id_value)
                except (ValueError, TypeError):
                    species_id = None
            else:
                species_id = None
            
            # 如果没有ID，使用自动递增
            if species_id is None:
                max_id += 1
                species_id = max_id
            else:
                max_id = max(species_id, max_id)
            
            species = {
                "id": species_id,
                "name": name,
                "code": code,
                "sequence": sequence,
                "length": len(sequence)
            }
            
            species_list.append(species)
            print(f"已读取: ID={species_id}, 名称={name}, 编号={code}, 序列长度={len(sequence)}")
        
        print(f"\n成功读取 {len(species_list)} 条参比序列")
        return species_list
        
    except Exception as e:
        print(f"读取Excel文件时出错: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def update_species_db(excel_file, output_file='species_db.json', backup=True):
    """更新参比序列数据库"""
    
    # 备份原文件
    if backup and os.path.exists(output_file):
        backup_file = output_file + '.backup'
        import shutil
        shutil.copy2(output_file, backup_file)
        print(f"已备份原文件到: {backup_file}")
    
    # 读取Excel
    species_list = read_excel_to_species_db(excel_file)
    
    if not species_list:
        print("错误: 无法从Excel文件读取数据")
        return False
    
    # 按ID排序
    species_list.sort(key=lambda x: x['id'])
    
    # 保存到JSON文件
    try:
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(species_list, f, ensure_ascii=False, indent=2)
        
        print(f"\n✅ 成功更新参比序列数据库!")
        print(f"   文件: {output_file}")
        print(f"   共 {len(species_list)} 条参比序列")
        print(f"   ID范围: {species_list[0]['id']} - {species_list[-1]['id']}")
        return True
        
    except Exception as e:
        print(f"保存JSON文件时出错: {str(e)}")
        return False

if __name__ == '__main__':
    excel_file = '参比序列-20260114.xlsx'
    
    if not os.path.exists(excel_file):
        print(f"错误: 找不到文件 {excel_file}")
        print("请确保Excel文件在当前目录下")
        sys.exit(1)
    
    print("=" * 60)
    print("参比序列数据库更新工具")
    print("=" * 60)
    print()
    
    success = update_species_db(excel_file)
    
    if success:
        print("\n更新完成！")
    else:
        print("\n更新失败！")
        sys.exit(1)
